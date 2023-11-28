# Django
from django.conf import settings
from django.http import JsonResponse
from django.conf import settings
from django.utils import timezone
from django.core.mail import send_mail
from django.db.models import Q

# Models, Forms
from .models import teacher, student

# Others
import os
import random
import string
import shutil
import base64
import smtplib
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
import hashlib
import openpyxl
import requests
from twilio.rest import Client
from datetime import datetime, timedelta


# Login
class user_login():
    username = ""
    authenticated = bool
    type = "None"
    user = teacher()
    
    def login(self, username, password):
        print("Checking for user")
        try:
            result = teacher.objects.filter(
                (Q(username = username) | Q(email = username) | Q(phone = username)) & 
                Q(password = hashlib.sha256(password.encode()).hexdigest())
            )
        except:
            result = student.objects.filter(
                (Q(username = username) | Q(email = username) | Q(phone = username)) & 
                Q(password = hashlib.sha256(password.encode()).hexdigest())
            )
        
        if result.exists():
            print(result)
            self.authenticated = True
            self.user = result.first()
            
            return self.authenticated, self.user
        else:
            return False, None
    
    
    def change_access_type(self, username):
        pass
    
    
    def get_user_details(self, username):
        try:
            result = teacher.objects.filter(Q(username = username))
        except:
            pass
    
    
    def if_user_exits(self, username):
        print("Checking for user")
        try:
            result = teacher.objects.filter(
                (Q(username = username) | Q(email = username) | Q(phone = username))
            )
        except:
            result = student.objects.filter(
                (Q(username = username) | Q(email = username) | Q(phone = username))
            )
        
        if result.exists():
            return True
        else:
            return False
    
    
    def change_user_password(username, password):
        pwd = hashlib.sha256(password.encode()).hexdigest()
        try:
            result = teacher.objects.filter(
                (Q(username = username) | Q(email = username) | Q(phone = username))
            )
        except:
            result = student.objects.filter(
                (Q(username = username) | Q(email = username) | Q(phone = username))
            )
        
        if result.exists():
            result[0].password = pwd
            result[0].save
            
            return True
        else:
            return False


# Xlsx extractor
class Xlsx_extractor():
    def get_file_extension(file_path):
        try:
            _, file_extension = os.path.splitext(file_path)
            return file_extension.lower() 
        except Exception as e:
            print(f"An error occurred: {e}")
            return None
        
        
    def file_upload(self, file):
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        type = self.get_file_extension(file)
        new_name = "teacher_details_{current_time}" + type
        
        try:
            os.rename(file, new_name)
            print(f"File renamed successfully to {new_name}")
        except FileNotFoundError:
            print("File not found.")
        except PermissionError:
            print("Permission error. Make sure you have the necessary permissions.")
        except Exception as e:
            print(f"An error occurred: {e}")
        
        try:
            shutil.move(file, "Teachers/")
            print(f"File moved successfully to {'Teachers/'}")
        except FileNotFoundError:
            print("File not found.")
        except PermissionError:
            print("Permission error. Make sure you have the necessary permissions.")
        except Exception as e:
            print(f"An error occurred: {e}")
        
        return "Teachers/", new_name, type, current_time
    
    
    def process_data(self, value):
        if isinstance(value, (int, float)):
            return int(value)
        
        return value
    

    def process_xlsx_teachers_data(self, data):
        temp = []

        for i in range(1, len(data)):
            x1 = data[i][3].split(",")
            str = data[i][4] 
            x2 = str[str.find("(")+1: -1].split("),")
            tt = {}

            for j in range(0, len(x1)):
                ff = x2[j][x2[j].find("(")+1 : len(x2[j])].split(",")
                ff = [int(i) for i in ff]

                tt[x1[j]] = ff

            temp.append({
                "Id" : data[i][0],
                "Name" : data[i][1],
                "Post" : data[i][2].split(", "),
                "Subjects" : tt,
                "Time Table" : [],
            })
            
        return temp


    def process_xlsx_subjects_data(self, data):
        temp = []
        clas = 32
        sec = "1"
        tt = []
        dat = {}
        
        for i in range(1, len(data)):
            try:
                if 0 < data[i][1] < 13 or data[i][1] == "Primary":
                    if clas != data[i][1]:
                        clas = data[i][1]
                        
                        if i != 1:
                            temp.append({
                                "Class" : data[i][1],
                                "Sections" : tt
                            })
                            tt = []

                    if sec != data[i][2]:
                        if i != 1:
                            tt.append(dat)
                        
                        dat = {}
                        dat["Section"] = data[i][2]
                        dat["Class Teacher"] = data[i][6]
                        dat["Subjects"] = []
                        dat["min"] = []
                        dat["max"] = []
                        sec = data[i][2]
                        clas = data[i][1]

                    if clas == data[i][1] and sec == data[i][2]:
                        dat["Subjects"].append(data[i][3])
                        dat["min"].append(data[i][4])
                        dat["max"].append(data[i][5])
                    
                    if i == len(data)-1:
                        tt.append(dat)
                        temp.append({
                            "Class" : data[i][1],
                            "Sections" : tt
                        })
            except:
                pass   

        return temp

    
    def extract_xlsx(self, file):
        # base_dir, filename, current_time = self.file_upload(file)
        # workbook = openpyxl.load_workbook(base_dir / filename)
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        max_row = sheet.max_row
        max_column = sheet.max_column
        data = []
        
        for row in range(1, max_row + 1):
            row_data = list(map(self.process_data, [sheet.cell(row = row, column = col).value for col in range(1, max_column + 1)]))
            data.append(row_data)
        
        return data

    
    def get_teachers_data(self, file):
        data = self.extract_xlsx(file)
        data = self.process_xlsx_teachers_data(data)

        return data
    

    def get_subject_list_data(self, file):
        data = self.extract_xlsx(file)
        data = self.process_xlsx_subjects_data(data)

        return data


# OTP
class OTP():
    ACCOUNT_SID = "ACd5a3b4236c548582b3079c4e4a8219b7"
    AUTH_TOKEN = "f7126fb59f26189bd21888869886c972"
    SERVICE_SID = ""
    TWILIO_PHONE_NUMBER = "+12512399796"
    SENDER_EMAIL = "vignesh20152@iiitd.ac.in"
    valid_till = datetime.now()
    otp = ""
    
    def generate_otp(self, length = 6):
        self.valid_till = datetime.now() + timedelta(minutes = 5)
        characters = string.digits
        self.otp = ''.join(random.choice(characters) for _ in range(length))
        
        return self.valid_till, self.otp


    def send_otp_phone(self, phone_number, otp):
        client = Client(self.ACCOUNT_SID, self.AUTH_TOKEN)
        message = client.messages.create(
            body = f'Your OTP is: {otp}',
            from_ = self.TWILIO_PHONE_NUMBER,
            to = phone_number
        )


    def send_otp_via_email(self, email, otp):
        HTML = '''
            <html>
                <head>
                </head>
                <body>
                    <h1>Your OTP: {}</h1>
                </body>
            </html>
        '''.format(otp)
        body = f'Your OTP is: {otp}'
        TO = [email]
        FROM = self.SENDER_EMAIL
        SUBJECT = "OTP for EduSchedula"
        CONTENT = ''
        
        MESSAGE = MIMEMultipart('related')
        part2 = MIMEText(HTML, 'html')
        MESSAGE['to'] = ', '.join(TO)
        MESSAGE['from'] = FROM
        MESSAGE['subject'] = SUBJECT
        MESSAGE.attach(part2)

        print('starting session')
        session = smtplib.SMTP('smtp.gmail.com')
        session.starttls()

        print('logging in')
        session.login('godtester04@gmail.com', 'eugcteapmxdglycu')
        print('logged in')

        TEXT = MESSAGE.as_bytes()

        print('sending email')
        session.sendmail(FROM, (TO), TEXT)
        session.quit()
        print('email sent')

        return True
    
    
    def verify_otp(self, otp):
        if datetime.now() <= self.valid_till:
            if otp == self.otp:
                return True, "Successful"
            else:
                return False, "Invalid OTP"
        else:
            return False, "OTP, time out. Try again"
    
    
    def initiate_verify_phone(self, phone):
        endpoint = f"https://verify.twilio.com/v2/Services/{self.SERVICE_SID}/Verifications"
        credentials = (self.ACCOUNT_SID, self.AUTH_TOKEN)
        
        data = {
            "To": phone,
            "Channel": "sms",
        }

        try:
            response = requests.post(endpoint, auth=credentials, json=data)

            if response.status_code in {200, 201}:
                print(f"Verification initiated for {phone}.")
                
                return response.json()["sid"] 
            else:
                print(f"Failed to initiate verification. Status code: {response.status_code}")
                print(response.text)
                
                return None

        except Exception as e:
            print(f"Error: {e}")
            return None
    
    
    def verify_phone(self, ver_sid, otp):
        endpoint = f"https://verify.twilio.com/v2/Services/{self.SERVICE_SID}/Verifications"
        credentials = (self.ACCOUNT_SID, self.AUTH_TOKEN)
        data = {"Code": otp}

        try:
            response = requests.post(endpoint, auth=credentials, json=data)

            if response.status_code == 200:
                print(f"Verification successful for SID {ver_sid}.")
            else:
                print(f"Verification failed. Status code: {response.status_code}")
                print(response.text)

        except Exception as e:
            print(f"Error: {e}")